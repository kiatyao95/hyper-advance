#!/usr/bin/env python3
"""Rebuild v4/public/data/catalog.json from References/ materials."""

from __future__ import annotations

import json
import re
import shutil
import unicodedata
from datetime import date
from pathlib import Path

import fitz
from openpyxl import load_workbook

try:
    from PIL import Image
except ImportError:
    Image = None

ROOT = Path(__file__).resolve().parents[2]
REFS = ROOT / "References"
DATA_DIR = ROOT / "27-06-2026"
LEGACY_DATA_DIR = ROOT / "23-06-2026"
PROJECT_DETAILS = LEGACY_DATA_DIR / "project details"
HA_MEDITECH_DIR = DATA_DIR / "HA MEDITECH PROJECT REFERENCE"
RESOURCES_DIR = DATA_DIR / "resources" / "resources"
PROJECT_PICS_DIR = RESOURCES_DIR / "project pics"
V4 = ROOT / "v4"
OUT = V4 / "public" / "data" / "catalog.json"
ASSETS = V4 / "public" / "assets"
BASE_CATALOG = OUT

LOCATIONS = {"KL", "SEL", "PJ", "PG", "JB", "SB", "WP", "Johor", "Sabah", "Penang", "Pahang", "KEDAH", "Kuala Lumpur", "Seremban", "Thailand", "Myanmar", "Singapore", "Saudi Arabia", "Kelantan", "Perak", "Putrajaya", "Melaka", "Pahang", "N Sembilan", "Vietnam", "Sri Lanka", "Kuwait", "Terengganu"}

SERIES_SET = {"GF", "GH", "GT", "IX", "IXG", "JO"}

TYPE_KEYWORDS = (
    "Audio Intercom", "B/W Video Intercom", "Color Video Intercom",
    "AUDIO INTERCOM", "AUDIO VIDEO INTERCOM", "Audio Video Intercom",
)

FEATURED_NAMES = {
    "four seasons hotel kl", "four seasons kl", "klcc", "hospital utar",
    "hospital universiti tunku abdul rahman", "k2 data centre", "pnb 118",
    "pnb office", "sunway medical centre damansara", "sunway medical centre velocity",
    "menara tan & tan", "midvalley centrepoint north & south", "gardens mall",
    "prince court medical centre", "istana terengganu", "genting premium outlet",
    "crowne plaza", "hilton kota kinabalu",
}

VERIFIED_AMOUNTS = {
    "four seasons hotel kl": "RM 4.4M",
    "four seasons kl": "RM 4.4M",
    "hospital utar": "RM 3.3M",
    "hospital universiti tunku abdul rahman": "RM 3.3M",
}

EXISTING_IMAGES = {
    "istana terengganu": "https://www.hyper-advance.com/assets/img/projects/Pernah%20Tengok%20Dalaman%20Istana%20Baru%20Diraja%20Terengganu.jpg",
    "prince court medical centre": "https://www.hyper-advance.com/assets/img/projects/2.jpg",
    "island hospital penang": "https://www.hyper-advance.com/assets/img/projects/5.jpg",
    "four seasons hotel kl": "https://www.hyper-advance.com/assets/img/projects/four-seasons-kuala-lumpur-malaysia-16.jpg",
    "eco tropics jb": "https://www.hyper-advance.com/assets/img/projects/7.jpg",
    "princess cove": "https://www.hyper-advance.com/assets/img/projects/8.jpg",
    "hk square": "https://www.hyper-advance.com/assets/img/projects/9.jpg",
    "sunway convention centre": "https://www.hyper-advance.com/assets/img/projects/11.jpg",
    "hosp sultan abdul halim": "https://www.hyper-advance.com/assets/img/projects/12.jpg",
    "mutiara ville cyberjaya": "https://www.hyper-advance.com/assets/img/projects/13.jpg",
    "laguna residences": "https://www.hyper-advance.com/assets/img/projects/14.jpg",
    "koi prima": "https://www.hyper-advance.com/assets/img/projects/15.jpg",
    "star residences klcc": "https://www.hyper-advance.com/assets/img/projects/17.jpg",
    "genting highland p outlet": "https://www.hyper-advance.com/assets/img/projects/18.jpg",
    "help university kl": "https://www.hyper-advance.com/assets/img/projects/19.jpg",
    "iskandar putri hospital": "https://www.hyper-advance.com/assets/img/projects/20.jpg",
    "kpj puteri": "https://www.hyper-advance.com/assets/img/projects/21.jpg",
    "ppum cic": "https://www.hyper-advance.com/assets/img/projects/22.jpg",
    "lai meng school": "https://www.hyper-advance.com/assets/img/projects/23.jpg",
    "kpj dato onn": "https://www.hyper-advance.com/assets/img/projects/24.jpg",
    "hospital cyberjaya": "https://www.hyper-advance.com/assets/img/projects/25.jpg",
    "hosp pengajar upm": "https://www.hyper-advance.com/assets/img/projects/26.jpg",
}

SYSTEM_TAG = {
    "nurse-call": "Nurse Call",
    "public-address": "Public Address",
    "av-intercom": "Intercom",
    "lighting-control": "Lighting Control",
    "smatv": "SMATV",
    "master-clock": "Master Clock",
    "ips": "IPS System",
    "card-access": "Card Access",
    "cctv": "CCTV",
    "audio-visual-system": "Audio Visual",
    "ot-tie-line": "OT Tie Line",
    "digital-call": "Digital Call",
    "fireman-intercom": "Fireman Intercom",
    "panic-button": "Panic Button",
    "intruder-alarm": "Intruder Alarm",
    "image-speak-through": "Image Speak Through",
}

BRAND_TO_DIST = {
    "austco": ("austco", "nurse-call"),
    "aiphone": ("aiphone", "av-intercom"),
    "ajb": ("ajb", "building-intercom"),
    "amperes": ("amperes", "public-address"),
    "lutron": ("lutron", "lighting-control"),
    "fagor": ("fagor", "smatv"),
    "bodet": ("bodet", "master-clock"),
    "esa grimma": ("esa-grimma", "ips"),
    "bsa grimma": ("esa-grimma", "ips"),
    "entrypass": ("amperes", "card-access"),
    "falco": ("amperes", "card-access"),
}

SYSTEM_PATTERNS = [
    (re.compile(r"nurse\s*call", re.I), "nurse-call", "austco", "Nurse Call"),
    (re.compile(r"intercom", re.I), "av-intercom", "aiphone", "Intercom"),
    (re.compile(r"public address|pa system|pa &", re.I), "public-address", "amperes", "Public Address"),
    (re.compile(r"smatv", re.I), "smatv", "fagor", "SMATV"),
    (re.compile(r"lighting control|lutron", re.I), "lighting-control", "lutron", "Lighting Control"),
    (re.compile(r"card acces|card access", re.I), "card-access", None, "Card Access"),
    (re.compile(r"cctv", re.I), "cctv", None, "CCTV"),
    (re.compile(r"audio visual|\bav system\b|\bav\b", re.I), "audio-visual-system", None, "Audio Visual"),
    (re.compile(r"ot tie line", re.I), "ot-tie-line", None, "OT Tie Line"),
    (re.compile(r"digital call", re.I), "digital-call", None, "Digital Call"),
    (re.compile(r"master clock", re.I), "master-clock", "bodet", "Master Clock"),
    (re.compile(r"fireman intercom", re.I), "fireman-intercom", None, "Fireman Intercom"),
    (re.compile(r"panic button", re.I), "panic-button", None, "Panic Button"),
    (re.compile(r"intruder alarm", re.I), "intruder-alarm", None, "Intruder Alarm"),
    (re.compile(r"image speak through", re.I), "image-speak-through", "amperes", "Image Speak Through"),
    (re.compile(r"conference|video conference", re.I), "audio-visual-system", None, "Audio Visual"),
]

SECTOR_MAP = {
    "healthcare environments": "healthcare",
    "commercial": "commercial",
    "hospitality": "hospitality",
    "residential": "residential",
}

EXCLUDED_SYSTEM_IDS = {"building-intercom", "ips"}

SYSTEM_OFFER_PROJECTS = {
    "nurse-call-system": [
        "sunway-medical-centre-velocity",
        "prince-court-medical-centre",
        "sunway-medical-centre-damansara",
        "hospital-tunku-ampuan-najihah",
    ],
    "intercom": [
        "klcc",
        "pjs-highway",
        "pnb-office",
        "k2-data-centre",
    ],
    "public-address": [
        "kl-wellness-city",
        "ioi-mall-damansara",
        "gardens-mall",
        "menara-tan-tan",
        "menara-hampshire",
    ],
    "smatv": [
        "princess-cove-phase-1",
        "kl-wellness-city",
        "hospital-jengka",
        "hospital-sultanah-maliha",
        "pusat-latihan-unit-tindakhas-semenyih",
    ],
    "lighting-control": [
        "w-hotel",
        "four-seasons-kl",
        "istana-syarqiyyah",
        "sunway-resort-hotel",
    ],
    "card-access": [
        "assunta-hospital",
        "menara-southpoint",
        "gardens-south-tower",
        "prince-court-medical-centre",
        "atwater-office",
    ],
    "cctv": [
        "masjid-bukit-rahman",
        "avant-warehouse",
        "lai-meng-school",
        "park-city-medical-centre",
        "prince-court-medical-centre",
    ],
    "audio-visual-system": [
        "cidb-office",
        "pusat-latihan-unit-tindakhas-semenyih",
    ],
}

CLIENT_LOGO_FILES = {
    "Sunway": "client logo/sunway healthcare.webp",
    "Park City Medical Centre": "client logo/park city.png",
    "KPJ": "client logo/kpj.jpg",
    "Gleneagles": "client logo/gleneagles.webp",
    "IOI": "client logo/ioi.svg",
    "KLCC": "client logo/klcc.png",
    "PNB": "client logo/pnb.png",
    "IGB": "client logo/igb.png",
    "JKR": "client logo/JKR_Malaysia_logo.png",
    "KKM": "client logo/kkm.jpg",
}


def norm(s: str) -> str:
    s = unicodedata.normalize("NFKD", s or "")
    s = s.encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"[^a-z0-9]+", " ", s.lower()).strip()
    return re.sub(r"\s+", " ", s)


def slugify(name: str, suffix: str = "") -> str:
    base = re.sub(r"[^a-z0-9]+", "-", norm(name))[:60].strip("-")
    return f"{base}-{suffix}" if suffix else base


def is_year(v: str) -> bool:
    return bool(re.fullmatch(r"20\d{2}", v.strip()))


def is_qty(v: str) -> bool:
    v = v.replace(",", "").strip()
    return v.isdigit() and int(v) < 50000


def is_location(v: str) -> bool:
    return v.strip() in LOCATIONS or v.strip().lower() in {x.lower() for x in LOCATIONS}


def is_series(v: str) -> bool:
    return v.strip().upper() in SERIES_SET


def is_type(v: str) -> bool:
    u = v.upper()
    return any(t.upper() in u for t in TYPE_KEYWORDS) or "INTERCOM" in u


def sector_for(name: str, category: str = "") -> str:
    n = norm(name)
    cat = category.upper()
    if "HOSPITAL" in n or "MEDICAL" in n or "KPJ" in n or "CLINIC" in n or "HEALTH" in n:
        return "healthcare"
    if "HOTEL" in n or "RESORT" in n or "HILTON" in n or "MARRIOTT" in n or "FOUR SEASON" in n:
        return "hospitality"
    if "CONDO" in n or "RESIDEN" in n or "VILLA" in n or "APARTMENT" in n or "TOWER" in n and "OFFICE" not in n:
        return "residential"
    if cat in ("HOTELS / RESORTS",):
        return "hospitality"
    if cat in ("EDUCATION & TRAINING", "SCHOOLS"):
        return "commercial"
    if cat in ("GOVERNMENT / PUBLIC", "SECURITY"):
        return "commercial"
    return "commercial"


def find_project_folder(name: str, folders: dict[str, Path]) -> Path | None:
    key = norm(name)
    if key in folders:
        return folders[key]
    for fn, path in folders.items():
        if key in fn or fn in key:
            return path
    return None


def copy_ha_meditech_images() -> list[dict]:
    """Copy HA Meditech IPS project photos; return project records for haMeditech section."""
    dest_root = ASSETS / "projects" / "ha-meditech"
    dest_root.mkdir(parents=True, exist_ok=True)
    projects: list[dict] = []

    if not HA_MEDITECH_DIR.exists():
        return projects

    for folder in sorted(HA_MEDITECH_DIR.iterdir()):
        if not folder.is_dir():
            continue

        slug = slugify(folder.name)
        dest_dir = dest_root / slug
        dest_dir.mkdir(parents=True, exist_ok=True)

        sources = sorted(
            [p for p in folder.iterdir() if p.is_file() and p.suffix.lower() in {".jpg", ".jpeg", ".png", ".webp"}],
            key=lambda p: (
                0 if p.stem.isdigit() else 1,
                int(p.stem) if p.stem.isdigit() else 0,
                p.stem.lower(),
            ),
        )

        paths: list[str] = []
        for i, src in enumerate(sources, 1):
            dest = dest_dir / f"{i}.jpg"
            if Image:
                with Image.open(src) as im:
                    im = im.convert("RGB")
                    w, h = im.size
                    if w > 1600:
                        im = im.resize((1600, int(h * 1600 / w)), Image.Resampling.LANCZOS)
                    im.save(dest, "JPEG", quality=82, optimize=True)
            else:
                shutil.copy2(src, dest)
            paths.append(f"/assets/projects/ha-meditech/{slug}/{i}.jpg")

        # Extract location from folder name (after comma)
        location = ""
        if "," in folder.name:
            location = folder.name.split(",", 1)[1].strip()

        display_name = folder.name.split(",")[0].strip().title()
        projects.append({
            "id": slug,
            "slug": slug,
            "name": display_name,
            "systemId": "ips",
            "distributorId": "esa-grimma",
            "tag": "IPS System",
            "location": location,
            "image": paths[0] if paths else "/assets/projects/placeholder-healthcare.jpg",
            "images": paths,
            "systemsCovered": ["IPS System"],
            "sector": "healthcare",
            "source": "ha-meditech",
        })

    return projects


def copy_client_logos() -> dict[str, str]:
    """Copy client logos from 27-06-2026 resources; return name -> public path."""
    dest_dir = ASSETS / "clients"
    dest_dir.mkdir(parents=True, exist_ok=True)
    logo_map: dict[str, str] = {}

    for name, rel in CLIENT_LOGO_FILES.items():
        src = RESOURCES_DIR / rel
        if not src.exists():
            continue
        ext = src.suffix.lower()
        dest_name = f"{slugify(name)}{ext}"
        dest = dest_dir / dest_name
        if ext in {".jpg", ".jpeg", ".png", ".webp"} and Image:
            with Image.open(src) as im:
                im = im.convert("RGBA" if im.mode in ("RGBA", "P") else "RGB")
                if im.mode == "RGBA":
                    im.save(dest, "PNG", optimize=True)
                else:
                    im.save(dest.with_suffix(".jpg"), "JPEG", quality=85, optimize=True)
                    dest = dest.with_suffix(".jpg")
        else:
            shutil.copy2(src, dest)
        logo_map[name] = f"/assets/clients/{dest.name}"

    return logo_map


def copy_supporting_brand_logos() -> dict[str, str]:
    """Copy supporting brand logos; return normalized brand token -> public path."""
    dest_dir = ASSETS / "logos" / "supporting"
    dest_dir.mkdir(parents=True, exist_ok=True)
    logo_map: dict[str, str] = {}

    brand_dir = RESOURCES_DIR / "brand logo"
    if not brand_dir.exists():
        return logo_map

    for src in brand_dir.iterdir():
        if not src.is_file():
            continue
        key = norm(src.stem)
        dest = dest_dir / src.name.lower().replace(" ", "-")
        shutil.copy2(src, dest)
        logo_map[key] = f"/assets/logos/supporting/{dest.name}"

    return logo_map


def brand_logos_for_text(brands_text: str, supporting: dict[str, str]) -> list[str]:
    logos: list[str] = []
    seen: set[str] = set()

    main_logos = {
        "aiphone": "/assets/logos/aiphone.png",
        "austco": "/assets/logos/austco.png",
        "amperes": "/assets/logos/amperes.png",
        "lutron": "/assets/logos/lutron.png",
        "fagor": "/assets/logos/fagor-brand.jpg",
        "bodet": "/assets/logos/bodet-header.svg",
        "zkteco": "/assets/logos/supporting/zkteco.png",
        "entrypass": "/assets/logos/supporting/entrypass.jpg",
        "bosch": "/assets/logos/supporting/bosch.png",
        "dahua": "/assets/logos/supporting/dahua.jpg",
        "hikvision": "/assets/logos/supporting/hikvision.jpg",
        "paradox": "/assets/logos/supporting/paradox.png",
        "yamaha": "/assets/logos/supporting/yamaha.png",
        "extron": "/assets/logos/supporting/extron.jpg",
        "aten": "/assets/logos/supporting/aten.png",
        "abtus": "/assets/logos/supporting/abtus.jpg",
        "amx": "/assets/logos/supporting/amx_logo.svg",
        "gms": "/assets/logos/supporting/gms.jpg",
        "myq": "/assets/logos/supporting/myq.jpg",
        "mictron": "/assets/logos/supporting/mictron.png",
    }

    for token in re.split(r"[,/&]+", brands_text or ""):
        key = norm(token.strip())
        if not key or key in seen:
            continue
        if key in main_logos:
            seen.add(key)
            logos.append(main_logos[key])
            continue
        for logo_key, path in supporting.items():
            if key in logo_key or logo_key in key:
                seen.add(key)
                logos.append(path)
                break
    return logos


def copy_folder_images(folder: Path, slug: str, projects_dir: Path) -> list[str]:
    """Copy image files from a folder into assets/projects/{slug}/."""
    dest_dir = projects_dir / slug
    dest_dir.mkdir(parents=True, exist_ok=True)

    sources = sorted(
        [p for p in folder.iterdir() if p.is_file() and p.suffix.lower() in {".jpg", ".jpeg", ".png", ".webp"}],
        key=lambda p: (
            0 if p.stem.isdigit() else 1,
            int(p.stem) if p.stem.isdigit() else 0,
            p.stem.lower(),
        ),
    )

    paths: list[str] = []
    for i, src in enumerate(sources, 1):
        dest = dest_dir / f"{i}.jpg"
        if Image:
            with Image.open(src) as im:
                im = im.convert("RGB")
                w, h = im.size
                if w > 1600:
                    im = im.resize((1600, int(h * 1600 / w)), Image.Resampling.LANCZOS)
                im.save(dest, "JPEG", quality=82, optimize=True)
        else:
            shutil.copy2(src, dest)
        paths.append(f"/assets/projects/{slug}/{i}.jpg")

    return paths


def copy_resource_project_images() -> dict[str, list[str]]:
    """Copy 27-06-2026/resources project pics; return norm(name) -> public paths."""
    projects_dir = ASSETS / "projects"
    projects_dir.mkdir(parents=True, exist_ok=True)
    image_map: dict[str, list[str]] = {}

    if not PROJECT_PICS_DIR.exists():
        return image_map

    for folder in PROJECT_PICS_DIR.iterdir():
        if not folder.is_dir():
            continue
        slug = slugify(folder.name)
        paths = copy_folder_images(folder, slug, projects_dir)
        if paths:
            image_map[norm(folder.name)] = paths

    return image_map


def copy_project_images() -> dict[str, list[str]]:
    """Copy 23-06-2026/project details images; return norm(name) -> public paths."""
    projects_dir = ASSETS / "projects"
    projects_dir.mkdir(parents=True, exist_ok=True)
    image_map: dict[str, list[str]] = {}

    if PROJECT_DETAILS.exists():
        folders = {norm(d.name): d for d in PROJECT_DETAILS.iterdir() if d.is_dir()}
        for folder_key, folder in folders.items():
            slug = slugify(folder.name)
            paths = copy_folder_images(folder, slug, projects_dir)
            if paths:
                image_map[folder_key] = paths

    # Supplement / override from 27-06-2026 resource project pics (e.g. KLCC)
    for folder_key, paths in copy_resource_project_images().items():
        image_map[folder_key] = paths

    return image_map


def resolve_image(name: str, sector: str, image_map: dict[str, list[str]] | None = None) -> str:
    key = norm(name)
    if image_map and key in image_map and image_map[key]:
        return image_map[key][0]
    if image_map:
        for fn, paths in image_map.items():
            if key in fn or fn in key:
                return paths[0]
    if key in EXISTING_IMAGES:
        return EXISTING_IMAGES[key]
    return f"/assets/projects/placeholder-{sector}.jpg"


def resolve_images(name: str, image_map: dict[str, list[str]] | None = None) -> list[str]:
    key = norm(name)
    if image_map and key in image_map:
        return image_map[key]
    if image_map:
        for fn, paths in image_map.items():
            if key in fn or fn in key:
                return paths
    return []


def parse_systems_text(text: str) -> list[tuple[str, str | None, str]]:
    """Return list of (systemId, distributorId, label) from systems column."""
    out: list[tuple[str, str | None, str]] = []
    seen: set[str] = set()
    for pattern, sid, did, label in SYSTEM_PATTERNS:
        if pattern.search(text or "") and sid not in seen:
            seen.add(sid)
            out.append((sid, did, label))
    return out


def sector_from_label(label: str) -> str:
    return SECTOR_MAP.get(norm(label), "commercial")


def normalize_completion(completion) -> str:
    if completion is None or str(completion).strip() == "":
        return "On-going / Maintenance"
    comp = str(completion).strip()
    if comp.lower() in ("on-going", "ongoing"):
        return "On-going"
    if isinstance(completion, (int, float)):
        return str(int(completion))
    return comp


def model_for_project(system_id: str, model: str) -> str:
    if not model or str(model).strip().lower() in ("none", ""):
        return ""
    if system_id in ("nurse-call", "av-intercom"):
        return str(model).strip()
    return ""


def project_record(
    name: str,
    *,
    system_id: str,
    distributor_id: str,
    model: str = "",
    sector: str = "commercial",
    completion: str = "",
    amount: str | None = None,
    unit_count: int | None = None,
    intercom_type: str = "",
    location: str = "",
    source: str = "",
    featured: bool = False,
    additional_systems: list[str] | None = None,
    systems_covered: list[str] | None = None,
    client: str = "",
    suffix: str = "",
    image_map: dict[str, list[str]] | None = None,
) -> dict:
    route_slug = slugify(name)
    sid = slugify(name, suffix) if suffix else route_slug
    images = resolve_images(name, image_map)
    rec = {
        "id": sid,
        "slug": route_slug,
        "name": name.strip(),
        "systemId": system_id,
        "distributorId": distributor_id,
        "model": model or "",
        "sector": sector,
        "tag": SYSTEM_TAG.get(system_id, system_id),
        "featured": featured,
        "image": resolve_image(name, sector, image_map),
        "source": source,
    }
    if images:
        rec["images"] = images
    if completion:
        rec["completionDate"] = completion if completion.lower() in ("on-going", "ongoing") else str(completion)
    if amount:
        rec["contractAmount"] = amount
    if unit_count:
        rec["unitCount"] = unit_count
    if intercom_type:
        rec["intercomType"] = intercom_type
    if location:
        rec["location"] = location
    if additional_systems:
        rec["additionalSystems"] = additional_systems
    if systems_covered:
        rec["systemsCovered"] = systems_covered
    if client:
        rec["client"] = client
    return rec


def parse_aiphone_q3() -> list[dict]:
    doc = fitz.open(REFS / "Aiphone Project References-Q32024.pdf")
    lines = [l.strip() for l in "".join(p.get_text() for p in doc).split("\n") if l.strip()]

    # skip headers until COMPLETED PROJECTS
    start = 0
    for i, l in enumerate(lines):
        if l == "COMPLETED PROJECTS":
            start = i + 1
            break

    projects = []
    i = start
    while i < len(lines):
        if not re.fullmatch(r"\d+(?:\.\d+)?", lines[i]):
            i += 1
            continue
        num = lines[i]
        i += 1
        if i >= len(lines):
            break
        name = lines[i]
        i += 1

        # optional location
        location = ""
        if i < len(lines) and is_location(lines[i]):
            location = lines[i]
            i += 1

        # may have second location line skipped
        series = ""
        if i < len(lines) and is_series(lines[i]):
            series = lines[i]
            i += 1

        qty = None
        if i < len(lines) and is_qty(lines[i]):
            qty = int(lines[i].replace(",", ""))
            i += 1

        intercom_type = ""
        if i < len(lines) and is_type(lines[i]):
            intercom_type = lines[i]
            i += 1

        year = ""
        if i < len(lines) and is_year(lines[i]):
            year = lines[i]
            i += 1

        # handle dual qty/type rows (e.g. ARA Bangsar)
        while i < len(lines) and is_qty(lines[i]):
            i += 1
            if i < len(lines) and is_type(lines[i]):
                intercom_type = lines[i] if not intercom_type else intercom_type
                i += 1
            if i < len(lines) and is_year(lines[i]):
                year = lines[i]
                i += 1

        if "Project Referenes" in name or name.startswith("Page "):
            continue

        n = norm(name)
        projects.append(project_record(
            name,
            system_id="av-intercom",
            distributor_id="aiphone",
            model=series or "GF/GH/GT",
            sector="residential",
            completion=year,
            unit_count=qty,
            intercom_type=intercom_type,
            location=location,
            source="aiphone-q3",
            featured=n in FEATURED_NAMES,
            suffix=num.replace(".", "-"),
        ))

    return projects


def parse_aiphone_q4() -> list[dict]:
    doc = fitz.open(REFS / "Aiphone Project References-Q42024(IX Series).pdf")
    lines = [l.strip() for l in "".join(p.get_text() for p in doc).split("\n") if l.strip()]

    projects = []
    i = 0
    while i < len(lines):
        if not re.fullmatch(r"\d+", lines[i]):
            i += 1
            continue
        i += 1
        if i >= len(lines):
            break
        name = lines[i]
        i += 1
        if name.startswith("For enquiry"):
            break
        location = lines[i] if i < len(lines) else ""
        i += 1
        series = lines[i] if i < len(lines) else ""
        i += 1
        qty = int(lines[i].replace(",", "")) if i < len(lines) and is_qty(lines[i]) else None
        i += 1
        intercom_type = lines[i] if i < len(lines) else ""
        i += 1
        year = lines[i] if i < len(lines) and is_year(lines[i]) else ""
        i += 1

        sec = sector_for(name)
        n = norm(name)
        projects.append(project_record(
            name.title() if name.isupper() else name,
            system_id="av-intercom",
            distributor_id="aiphone",
            model=series,
            sector=sec,
            completion=year,
            unit_count=qty,
            intercom_type=intercom_type,
            location=location,
            source="aiphone-q4",
            featured=n in FEATURED_NAMES,
        ))
    return projects


def parse_bodet() -> list[dict]:
    doc = fitz.open(REFS / "BODET INSTALLATION BY HYPER ADVANCE.pdf")
    text = "".join(p.get_text() for p in doc)
    rows = re.findall(
        r"(\d+)\s+(.+?)\s+(20\d{2})\s+([\d,]+\.?\d*)",
        text,
    )
    projects = []
    for _, title, year, amount in rows:
        amt = f"RM {amount.strip().replace(',', '')}"
        if "." in amt.split()[-1]:
            val = float(amt.split()[-1])
            amt = f"RM {int(val):,}" if val >= 1000 else f"RM {val:,.0f}"
        n = norm(title)
        projects.append(project_record(
            title.strip(),
            system_id="master-clock",
            distributor_id="bodet",
            model="Master Clock",
            sector=sector_for(title),
            completion=year,
            amount=amt,
            source="bodet-pdf",
            featured=n in FEATURED_NAMES,
        ))
    return projects


def parse_amperes() -> list[dict]:
    doc = fitz.open(REFS / "Amperes_Ref_2023_compressed.pdf")
    lines = [l.strip() for l in "".join(p.get_text() for p in doc).split("\n") if l.strip()]

    projects = []
    i = 0
    while i < len(lines):
        line = lines[i]
        if line in (
            "Project References", "CATEGORY", "INSTALLATION", "LOCATION", "YEAR",
            "Providing dedicated Public Address System since 1999",
            "amperes electronics sdn bhd", "www.ampereselectronics.com",
            "Updated as at Nov 2020", "2014 - 2023",
        ) or line.startswith("The list comprises") or line.startswith("Source:") or line.startswith("Older references"):
            i += 1
            continue

        if line.isupper() and len(line) > 8 and "/" in line or line in (
            "OFFICE TOWERS", "HOTELS / RESORTS", "EDUCATION & TRAINING", "SCHOOLS",
            "ENTERTAINMENT & SHOPPING, COMMERCIAL, MIXED DEV.", "SECURITY",
            "GOVERNMENT / PUBLIC",
        ):
            category = line
            i += 1
            installs = []
            while i < len(lines):
                if lines[i].isupper() and len(lines[i]) > 10:
                    break
                if is_year(lines[i]) or lines[i] in LOCATIONS or lines[i].lower() in {x.lower() for x in LOCATIONS}:
                    break
                if lines[i] in ("Project References", "CATEGORY"):
                    break
                installs.append(lines[i])
                i += 1
            locations = []
            while i < len(lines):
                if is_year(lines[i]):
                    break
                if lines[i].isupper() and len(lines[i]) > 10 and "/" in lines[i]:
                    break
                locations.append(lines[i])
                i += 1
            years = []
            while i < len(lines):
                if not is_year(lines[i]) and lines[i] not in ("2016 / 17",):
                    break
                y = lines[i].replace(" / 17", "").strip()
                if is_year(y):
                    years.append(y)
                i += 1
                if len(years) >= len(installs):
                    break

            count = min(len(installs), len(years))
            for j in range(count):
                name = installs[j]
                loc = locations[j] if j < len(locations) else ""
                year = years[j]
                n = norm(name)
                projects.append(project_record(
                    name,
                    system_id="public-address",
                    distributor_id="amperes",
                    model="PA System",
                    sector=sector_for(name, category),
                    completion=year,
                    location=loc,
                    source="amperes-pdf",
                    featured=n in FEATURED_NAMES,
                    suffix=f"{slugify(category)[:20]}-{j}",
                ))
            continue
        i += 1
    return projects


def parse_xlsx(image_map: dict[str, list[str]] | None = None) -> list[dict]:
    xlsx_path = DATA_DIR / "Project Details (1).xlsx"
    if not xlsx_path.exists():
        xlsx_path = DATA_DIR / "Project Details.xlsx"
    if not xlsx_path.exists():
        xlsx_path = LEGACY_DATA_DIR / "Project Details.xlsx"
    if not xlsx_path.exists():
        xlsx_path = REFS / "Project Details.xlsx"

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["Project reference"]
    projects = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue

        name = str(row[0]).strip()
        client = str(row[1] or "").strip()
        sector_label = str(row[2] or "").strip()
        model = str(row[3] or "").strip()
        brand = str(row[4] or "").strip().lower()
        systems_text = str(row[5] or "")
        completion = row[6]

        parsed = parse_systems_text(systems_text)
        brand_key = brand.replace(" ", "")
        dist_id, sys_id = BRAND_TO_DIST.get(brand_key, (None, None))
        if not dist_id and parsed:
            sys_id, dist_id, _ = parsed[0]
            dist_id = dist_id or "amperes"
            sys_id = sys_id or "public-address"
        elif not dist_id:
            sys_id, dist_id = "public-address", "amperes"

        if sys_id in EXCLUDED_SYSTEM_IDS:
            if parsed:
                for sid, did, _ in parsed:
                    if sid not in EXCLUDED_SYSTEM_IDS:
                        sys_id, dist_id = sid, did or dist_id
                        break

        additional = [s[0] for s in parsed if s[0] != sys_id]
        systems_covered = [s[2] for s in parsed] or [SYSTEM_TAG.get(sys_id, sys_id)]
        sector = sector_from_label(sector_label) if sector_label else sector_for(name)
        n = norm(name)
        completion_str = normalize_completion(completion)
        model_str = model_for_project(sys_id, model)

        projects.append(project_record(
            name.title() if name.isupper() else name,
            system_id=sys_id,
            distributor_id=dist_id,
            model=model_str,
            sector=sector,
            completion=completion_str,
            source="xlsx",
            featured=n in FEATURED_NAMES,
            additional_systems=additional or None,
            systems_covered=systems_covered,
            client=client,
            image_map=image_map,
        ))

        for sid, did, _ in parsed:
            if sid == sys_id or sid in EXCLUDED_SYSTEM_IDS:
                continue
            projects.append(project_record(
                name.title() if name.isupper() else name,
                system_id=sid,
                distributor_id=did or dist_id,
                model=model_for_project(sid, model),
                sector=sector,
                completion=completion_str,
                source="xlsx-secondary",
                featured=False,
                systems_covered=systems_covered,
                client=client,
                suffix=sid,
                image_map=image_map,
            ))

    return projects


def merge_projects(sources: list[list[dict]]) -> list[dict]:
    merged: dict[str, dict] = {}
    order: list[str] = []

    def key(p: dict) -> str:
        if p.get("source") in ("aiphone-q3", "amperes-pdf"):
            return p["id"]
        return f"{norm(p['name'])}|{p['distributorId']}|{p.get('systemId','')}|{p.get('completionDate','')}|{p.get('unitCount','')}"

    for batch in sources:
        for p in batch:
            k = key(p)
            if k not in merged:
                merged[k] = p
                order.append(k)
            else:
                existing = merged[k]
                for field in ("model", "completionDate", "contractAmount", "unitCount", "intercomType", "location", "additionalSystems", "image", "images", "slug"):
                    if p.get(field) and not existing.get(field):
                        existing[field] = p[field]
                if p.get("featured"):
                    existing["featured"] = True
                if p.get("source") == "xlsx":
                    existing["source"] = "xlsx"

    # featured overrides (no contract amounts)
    for p in merged.values():
        p.pop("contractAmount", None)
        n = norm(p["name"])
        if n in FEATURED_NAMES:
            p["featured"] = True
        if n in ("four seasons kl", "four seasons hotel kl"):
            p["completionDate"] = "2017"
            p["featured"] = True
        if n == "klcc":
            p["completionDate"] = "2023"
            p["featured"] = True

    return [merged[k] for k in order]


def apply_local_images(projects: list[dict], image_map: dict[str, list[str]]) -> None:
    for p in projects:
        if "slug" not in p:
            p["slug"] = slugify(p["name"])
        images = resolve_images(p["name"], image_map)
        if images:
            p["images"] = images
            p["image"] = images[0]


def build_industry_projects(projects: list[dict]) -> dict:
    sectors = ["healthcare", "hospitality", "commercial", "residential"]
    out: dict[str, list] = {s: [] for s in sectors}
    seen: dict[str, set[str]] = {s: set() for s in sectors}

    dist_by_system: dict[str, str] = {}
    for _, sid, did, _ in SYSTEM_PATTERNS:
        if did:
            dist_by_system[sid] = did
    for _, (did, sid) in BRAND_TO_DIST.items():
        dist_by_system[sid] = did

    primary = [p for p in projects if p.get("source") == "xlsx"]
    sorted_projects = sorted(primary, key=lambda p: (not p.get("featured"), p["name"]))

    for p in sorted_projects:
        sec = p.get("sector", "commercial")
        if sec not in out:
            sec = "commercial"
        nkey = norm(p["name"])
        if nkey in seen[sec] or len(out[sec]) >= 15:
            continue

        links = [{"systemId": p["systemId"], "distributorId": p["distributorId"]}]
        for sid in p.get("additionalSystems") or []:
            if sid in EXCLUDED_SYSTEM_IDS:
                continue
            did = dist_by_system.get(sid) or p["distributorId"]
            if not any(l["systemId"] == sid for l in links):
                links.append({"systemId": sid, "distributorId": did})

        out[sec].append({"name": p["name"], "links": links})
        seen[sec].add(nkey)

    return out


def parse_systems_we_offer(supporting_logos: dict[str, str]) -> list[dict]:
    """Build systems-we-offer list from Excel with curated project slugs."""
    xlsx_path = DATA_DIR / "Project Details (1).xlsx"
    if not xlsx_path.exists():
        xlsx_path = DATA_DIR / "Project Details.xlsx"
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["Systems we offer "]

    offers = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        name = str(row[0]).strip()
        brands = str(row[1] or "").strip()
        description = str(row[2] or "").strip()
        if not description:
            continue

        offer_id = slugify(name)
        matched_slugs = SYSTEM_OFFER_PROJECTS.get(offer_id, [])

        # Intercom brands: Aiphone only (remove AJB per notes)
        if offer_id == "intercom":
            brands = "Aiphone"

        offers.append({
            "id": offer_id,
            "name": name.strip().rstrip(","),
            "brands": brands,
            "description": description,
            "projectSlugs": matched_slugs,
            "brandLogos": brand_logos_for_text(brands, supporting_logos),
        })

    return offers


CLIENT_REFERENCES = [
    {"name": "Sunway", "projectSlugs": ["sunway-resort-hotel", "sunway-medical-centre-damansara", "sunway-medical-centre-velocity"]},
    {"name": "Park City Medical Centre", "projectSlugs": ["park-city-medical-centre"]},
    {"name": "KPJ", "projectSlugs": ["kpj-ampang", "kpj-rawang"]},
    {"name": "Gleneagles", "projectSlugs": []},
    {"name": "IOI", "projectSlugs": ["ioi-mall-damansara"]},
    {"name": "KLCC", "projectSlugs": ["klcc"]},
    {"name": "PNB", "projectSlugs": ["pnb-office"]},
    {"name": "IGB", "projectSlugs": ["gardens-mall", "menara-tan-tan", "menara-hampshire", "menara-southpoint", "gardens-south-tower", "midvalley-centrepoint-north-south"]},
    {"name": "JKR", "projectSlugs": []},
    {"name": "KKM", "projectSlugs": ["hospital-tunku-ampuan-najihah", "hospital-seri-manjung", "hospital-baling"]},
]

AUTHORIZED_BRAND_IDS = ["aiphone", "austco", "amperes", "lutron", "fagor"]

SERVICE_CAPABILITIES = [
    {
        "id": "design",
        "icon": "fa-drafting-compass",
        "title": "Design",
        "description": "Design proposals for ELV systems, including layout plans and schematic drawings aligned to your project requirements.",
    },
    {
        "id": "supply",
        "icon": "fa-truck",
        "title": "Supply",
        "description": "A wide range of ELV products and systems from authorised brands and established partners.",
    },
    {
        "id": "installation",
        "icon": "fa-screwdriver-wrench",
        "title": "Installation",
        "description": "Capability to install virtually all types of ELV systems across commercial, residential, healthcare, and hospitality sites.",
    },
    {
        "id": "maintenance",
        "icon": "fa-wrench",
        "title": "Maintenance",
        "description": "Comprehensive and non-comprehensive maintenance programmes to keep every ELV system running smoothly.",
    },
    {
        "id": "programming",
        "icon": "fa-microchip",
        "title": "Programming",
        "description": "System programming configured to end-user needs, operational preferences, and site-specific workflows.",
    },
    {
        "id": "testing",
        "icon": "fa-flask",
        "title": "Testing & Commissioning",
        "description": "Structured testing, commissioning, and handover to ensure systems are ready for operational use.",
    },
    {
        "id": "training",
        "icon": "fa-graduation-cap",
        "title": "Training",
        "description": "User training on system operation, routine care, and best practices for long-term reliability.",
    },
]

HA_MEDITECH = {
    "title": "HA Meditech Sdn Bhd",
    "description": (
        "Sister company of Hyper Advance specialising solely in IPS systems for hospitals. "
        "Sole distributor of Esa Grimma — a trusted German brand. "
        "Led by an experienced engineer with more than 30 years in the field. Established in 2011."
    ),
    "systems": [
        {
            "name": "Isolated Power Supply (IPS)",
            "brands": "Esa Grimma",
            "distributorId": "esa-grimma",
            "systemId": "ips",
            "description": "Medical-grade isolated power systems for operating theatres, critical care zones, and hospital infrastructure.",
        },
    ],
    "projects": [],
}

FAGOR_MERGER_NOTE = (
    "Fagor SMATV systems distribute satellite and terrestrial TV signals across hotels, apartments, and commercial developments. "
    "The range also carries forward the engineering legacy of Ikusi — an established name in professional TV distribution — "
    "following its integration under the Fagor Group, combining proven expertise with expanded product support under one brand."
)


def enrich_distributors(distributors: list[dict]) -> list[dict]:
    out = []
    for d in distributors:
        rec = dict(d)
        if rec["id"] == "fagor":
            rec["logo"] = "/assets/logos/fagor-brand.jpg"
            rec.pop("logoAlt", None)
            rec["description"] = FAGOR_MERGER_NOTE
            rec["mergerNote"] = "Ikusi is now part of the Fagor Group. Hyper Advance supports both Fagor and legacy Ikusi SMATV deployments."
            rec["legacyImage"] = "/assets/logos/fagor.jpg"
        out.append(rec)
    return out


def copy_assets() -> list[str]:
    office_dir = ASSETS / "office"
    refs_dir = ASSETS / "references"
    projects_dir = ASSETS / "projects"
    brand_dir = ASSETS / "brand"
    for d in (office_dir, refs_dir, projects_dir, brand_dir):
        d.mkdir(parents=True, exist_ok=True)

    # reference PDFs
    pdf_map = {
        "aiphone-q32024.pdf": "Aiphone Project References-Q32024.pdf",
        "aiphone-q42024.pdf": "Aiphone Project References-Q42024(IX Series).pdf",
        "amperes-2023.pdf": "Amperes_Ref_2023_compressed.pdf",
        "bodet-installations.pdf": "BODET INSTALLATION BY HYPER ADVANCE.pdf",
    }
    for dest, src in pdf_map.items():
        shutil.copy2(REFS / src, refs_dir / dest)

    # logo
    logo_src = REFS / "HYPER TRANSPARENT LOGO WITH NAME.png"
    if logo_src.exists():
        shutil.copy2(logo_src, brand_dir / "hyper-advance-logo.png")

    fagor_src = RESOURCES_DIR / "brand logo" / "fagor.jpg"
    if fagor_src.exists():
        shutil.copy2(fagor_src, ASSETS / "logos" / "fagor-brand.jpg")

    # office photos
    gallery = []
    src_dir = REFS / "Hyper Advance Office" / "Hyper Advance Office Photos"
    if src_dir.exists():
        for img_path in sorted(src_dir.glob("*.jpg")):
            dest = office_dir / img_path.name
            if Image:
                with Image.open(img_path) as im:
                    im = im.convert("RGB")
                    w, h = im.size
                    if w > 1400:
                        im = im.resize((1400, int(h * 1400 / w)), Image.Resampling.LANCZOS)
                    im.save(dest, "JPEG", quality=82, optimize=True)
            else:
                shutil.copy2(img_path, dest)
            gallery.append(f"/assets/office/{img_path.name}")

    # placeholder images per sector
    sectors = ["healthcare", "hospitality", "commercial", "residential"]
    placeholder_colors = {
        "healthcare": (30, 80, 140),
        "hospitality": (120, 90, 40),
        "commercial": (50, 50, 80),
        "residential": (60, 100, 90),
    }
    if Image:
        for sec in sectors:
            path = projects_dir / f"placeholder-{sec}.jpg"
            if not path.exists():
                img = Image.new("RGB", (800, 500), placeholder_colors[sec])
                img.save(path, "JPEG", quality=85)

    return gallery


def main():
    with open(BASE_CATALOG, encoding="utf-8") as f:
        base = json.load(f)

    gallery = copy_assets()
    image_map = copy_project_images()
    ha_meditech_projects = copy_ha_meditech_images()
    client_logos = copy_client_logos()
    supporting_logos = copy_supporting_brand_logos()

    xlsx = parse_xlsx(image_map)
    projects = merge_projects([xlsx])
    apply_local_images(projects, image_map)
    systems_we_offer = parse_systems_we_offer(supporting_logos)
    distributors = enrich_distributors(base["distributors"])
    systems = [s for s in base["systems"] if s["id"] not in EXCLUDED_SYSTEM_IDS]

    client_refs = []
    for ref in CLIENT_REFERENCES:
        rec = dict(ref)
        if ref["name"] in client_logos:
            rec["logo"] = client_logos[ref["name"]]
        client_refs.append(rec)

    ha_meditech = dict(HA_MEDITECH)
    ha_meditech["projects"] = ha_meditech_projects

    # ensure unique ids
    used_ids: set[str] = set()
    for p in projects:
        base_id = p["id"]
        n = 1
        while p["id"] in used_ids:
            p["id"] = f"{base_id}-{n}"
            n += 1
        used_ids.add(p["id"])

    catalog = {
        "meta": {
            "company": "Hyper Advance Sdn Bhd",
            "version": "2.1",
            "updated": date.today().isoformat(),
            "projectCount": len(projects),
        },
        "systems": systems,
        "distributors": distributors,
        "authorizedBrandIds": AUTHORIZED_BRAND_IDS,
        "clientReferences": client_refs,
        "systemsWeOffer": systems_we_offer,
        "serviceCapabilities": SERVICE_CAPABILITIES,
        "haMeditech": ha_meditech,
        "keyDistributorship": [
            {"id": "intercom", "title": "INTERCOM SYSTEM", "icon": "fa-video", "brandIds": ["aiphone"]},
            {"id": "nurse-call", "title": "NURSE CALL SYSTEM", "icon": "fa-user-nurse", "brandIds": ["austco"]},
            {"id": "pa", "title": "PA SYSTEM", "icon": "fa-bullhorn", "brandIds": ["amperes"]},
            {"id": "lighting", "title": "LIGHTING CONTROL SYSTEM", "icon": "fa-lightbulb", "brandIds": ["lutron"]},
            {"id": "smatv", "title": "SMATV SYSTEM", "icon": "fa-tv", "brandIds": ["fagor"]},
        ],
        "company": {
            "founded": 1995,
            "staffCount": 30,
            "systemsOffered": 10,
            "authorizedBrandCount": 5,
            "yearsExperience": date.today().year - 1995,
            "successStories": [
                {
                    "name": "Four Seasons Hotel Kuala Lumpur",
                    "system": "Lutron Lighting & Curtain Control",
                    "completionDate": "2017",
                    "detail": "209 luxurious guest rooms",
                    "distributorId": "lutron",
                    "systemId": "lighting-control",
                },
                {
                    "name": "KLCC Twin Towers",
                    "system": "Aiphone IX Intercom",
                    "completionDate": "2023",
                    "detail": "280 IP video door stations for emergency access across both towers",
                    "distributorId": "aiphone",
                    "systemId": "av-intercom",
                },
                {
                    "name": "Hospital UTAR",
                    "system": "Full ELV — Nurse Call, CCTV, Card Access, PA",
                    "completionDate": "2022",
                    "detail": "250-bed teaching hospital",
                    "distributorId": "austco",
                    "systemId": "nurse-call",
                },
            ],
            "hotelReferences": [
                {"name": "St Giles Kuala Lumpur", "systems": "Card Access, PA, CCTV"},
                {"name": "Hard Rock Penang", "systems": "Card Access, PA, CCTV"},
                {"name": "Furama Hotel Kuala Lumpur", "systems": "Card Access, PA, CCTV, Barrier Gate"},
                {"name": "Boulevard Hotel Mid Valley", "systems": "Card Access, PA, CCTV, Barrier Gate"},
                {"name": "Sheraton Petaling Jaya", "systems": "Lighting Control System"},
                {"name": "Hilton Garden Inn Puchong", "systems": "Lighting Control System"},
                {"name": "Hotel Seri Iskandar", "systems": "Hotel Lockset System"},
                {"name": "Ombak Villa Langkawi", "systems": "Hotel Lockset System"},
            ],
            "officeGallery": gallery,
            "supportingBrands": [
                {"system": "Fireman Intercom", "brands": "Mictron"},
                {"system": "Digital Call", "brands": "GMS, MYQ"},
                {"system": "Master Clock", "brands": "National Time, Bodet"},
                {"system": "Conference", "brands": "Bosch"},
                {"system": "Panic Button", "brands": "Austco, Paradox"},
                {"system": "Intruder Alarm", "brands": "Paradox"},
                {"system": "Card Access", "brands": "ZKTeco, EntryPass, Microengine, Bosch, Dahua, HIKVision"},
                {"system": "CCTV", "brands": "Dahua, HIKVision, Bosch"},
                {"system": "Audio Visual", "brands": "AMX, Abtus, Yamaha, Aten, Extron"},
            ],
        },
        "referenceDocuments": [
            {"label": "Aiphone Residential References (Q3 2024)", "url": "/assets/references/aiphone-q32024.pdf", "distributorId": "aiphone"},
            {"label": "Aiphone Non-Residential IX Series (Q4 2024)", "url": "/assets/references/aiphone-q42024.pdf", "distributorId": "aiphone"},
            {"label": "Amperes Project References (2023)", "url": "/assets/references/amperes-2023.pdf", "distributorId": "amperes"},
            {"label": "Bodet Installations by Hyper Advance", "url": "/assets/references/bodet-installations.pdf", "distributorId": "bodet"},
        ],
        "projects": projects,
        "industryProjects": build_industry_projects(projects),
    }

    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(catalog, f, indent=2, ensure_ascii=False)
        f.write("\n")

    print(f"Wrote {len(projects)} projects to {OUT}")
    print(f"Office gallery: {len(gallery)} images")
    print(f"Project image folders: {len(image_map)}")
    print(f"HA Meditech IPS projects: {len(ha_meditech_projects)}")
    print(f"Client logos: {len(client_logos)}")


if __name__ == "__main__":
    main()
